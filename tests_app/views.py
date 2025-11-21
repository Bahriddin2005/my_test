from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.utils import timezone
from django.db import transaction
from django.core.paginator import Paginator
import json
import random
from .models import Test, Question, Choice, TestAttempt, Answer, TestResult, TestRetakeRequest
from accounts.models import User
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

@login_required
@require_http_methods(["POST"])
def pause_test(request, test_id):
    """Testni pauza qilish - Admin uchun"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    try:
        test = get_object_or_404(Test, id=test_id)
        # Migration qo'llanmagan bo'lsa ham ishlashi uchun
        if hasattr(test, 'is_paused'):
            test.is_paused = True
            if hasattr(test, 'paused_at'):
                test.paused_at = timezone.now()
            test.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Test pauza qilindi',
                'is_paused': True,
                'paused_at': test.paused_at.isoformat() if hasattr(test, 'paused_at') and test.paused_at else None
            })
        else:
            return JsonResponse({
                'error': 'Migration qo\'llanmagan. Migrationni ishga tushiring: python manage.py migrate tests_app'
            }, status=500)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Error in pause_test: {str(e)}', exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@require_http_methods(["POST"])
def resume_test(request, test_id):
    """Testni davom ettirish - Admin uchun"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    try:
        test = get_object_or_404(Test, id=test_id)
        # Migration qo'llanmagan bo'lsa ham ishlashi uchun
        if hasattr(test, 'is_paused'):
            test.is_paused = False
            if hasattr(test, 'paused_at'):
                test.paused_at = None
            test.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Test davom ettirildi',
                'is_paused': False
            })
        else:
            return JsonResponse({
                'error': 'Migration qo\'llanmagan. Migrationni ishga tushiring: python manage.py migrate tests_app'
            }, status=500)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Error in resume_test: {str(e)}', exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def test_control_view(request, test_id):
    """Test nazorat sahifasi - Admin uchun"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    test = get_object_or_404(Test, id=test_id)
    
    # Test yechayotgan o'quvchilar
    active_attempts = TestAttempt.objects.filter(
        test=test,
        is_completed=False
    ).select_related('student').order_by('-started_at')
    
    attempts_data = []
    for attempt in active_attempts:
        elapsed_time = timezone.now() - attempt.started_at
        attempts_data.append({
            'id': attempt.id,
            'student_name': attempt.student.get_full_name() or attempt.student.username,
            'student_username': attempt.student.username,
            'student_grade': attempt.student.grade,
            'student_class': attempt.student.class_name,
            'started_at': attempt.started_at.isoformat(),
            'elapsed_time': str(elapsed_time).split('.')[0] if elapsed_time else '0:00:00'
        })
    
    context = {
        'test': test,
        'active_attempts': attempts_data,
        'total_active': len(attempts_data)
    }
    
    if request.headers.get('Accept') == 'application/json':
        # Migration qo'llanmagan bo'lsa ham ishlashi uchun
        is_paused = getattr(test, 'is_paused', False)
        paused_at = None
        if hasattr(test, 'paused_at') and test.paused_at:
            paused_at = test.paused_at.isoformat()
        
        return JsonResponse({
            'test': {
                'id': test.id,
                'title': test.title,
                'is_paused': is_paused,
                'paused_at': paused_at,
                'is_active': test.is_active
            },
            'active_attempts': attempts_data,
            'total_active': len(attempts_data)
        })
    
    return render(request, 'tests_app/test_control.html', context)

@login_required
def test_time_view(request, test_id):
    """Server vaqtini qaytarish - Test vaqtini hisoblash uchun"""
    test = get_object_or_404(Test, id=test_id)
    
    # O'quvchi faqat o'z testi uchun vaqtni olishi mumkin
    if request.user.role == 'student':
        if test.grade != request.user.grade:
            return JsonResponse({'error': 'Access denied'}, status=403)
    
    # Migration qo'llanmagan bo'lsa ham ishlashi uchun
    is_paused = getattr(test, 'is_paused', False)
    
    return JsonResponse({
        'server_time': timezone.now().isoformat(),
        'test_id': test.id,
        'is_paused': is_paused
    })

@login_required
def monitor_view(request):
    """Test monitoring sahifasi - Admin uchun barcha testlarni nazorat qilish"""
    if request.user.role != 'admin':
        return redirect('accounts:dashboard')
    
    if request.method == 'GET' and request.headers.get('Accept') == 'application/json':
        from django.db import connection
        
        try:
            # Raw SQL yordamida test ID'larni olish
            test_ids = []
            try:
                with connection.cursor() as cursor:
                    cursor.execute("SELECT id FROM tests_app_test ORDER BY created_at DESC")
                    test_ids = [row[0] for row in cursor.fetchall()]
            except Exception as db_error:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f'Error fetching test IDs in monitor_view: {str(db_error)}')
                test_ids = []
            
            # Test obyektlarini xavfsiz yuklash
            tests = []
            for test_id in test_ids:
                try:
                    test = Test.objects.select_related('created_by').get(id=test_id)
                    tests.append(test)
                except Exception:
                    continue
            
            tests_data = []
            for test in tests:
                try:
                    # Faol urinishlar (yechilayotgan testlar)
                    active_attempts = TestAttempt.objects.filter(
                        test=test,
                        is_completed=False
                    ).count()
                    
                    # Tugallangan urinishlar
                    completed_attempts = TestAttempt.objects.filter(
                        test=test,
                        is_completed=True
                    ).count()
                    
                    # Migration qo'llanmagan bo'lsa ham ishlashi uchun
                    is_paused = getattr(test, 'is_paused', False)
                    paused_at = None
                    if hasattr(test, 'paused_at') and test.paused_at:
                        paused_at = test.paused_at.isoformat()
                    
                    tests_data.append({
                        'id': test.id,
                        'title': test.title,
                        'subject': test.subject,
                        'grade': test.grade,
                        'is_active': test.is_active,
                        'is_paused': is_paused,
                        'paused_at': paused_at,
                        'active_attempts': active_attempts,
                        'completed_attempts': completed_attempts,
                        'total_questions': test.total_questions,
                        'time_limit': test.time_limit,
                        'created_by': test.created_by.get_full_name() if test.created_by and hasattr(test.created_by, 'get_full_name') else (test.created_by.username if test.created_by else 'Noma\'lum'),
                        'created_at': test.created_at.isoformat() if test.created_at else '',
                    })
                except Exception as e:
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.error(f'Error processing test {test.id} in monitor_view: {str(e)}')
                    continue
            
            return JsonResponse({
                'tests': tests_data,
                'total_tests': len(tests_data)
            })
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f'Error in monitor_view: {str(e)}', exc_info=True)
            return JsonResponse({
                'error': 'Ma\'lumotlarni yuklashda xatolik yuz berdi',
                'detail': str(e)
            }, status=500)
    
    return render(request, 'tests_app/monitor.html')

@login_required
def test_list_view(request):
    """List all available tests for students or created tests for teachers"""
    if request.method == 'GET' and request.headers.get('Accept') == 'application/json':
        from django.db import connection
        
        if request.user.role == 'student':
            try:
                # Raw SQL yordamida test ID'larni olish (is_paused maydonini tekshirmaslik uchun)
                test_ids = []
                try:
                    # Student'ning grade maydonini tekshirish
                    student_grade = getattr(request.user, 'grade', None)
                    if student_grade is None:
                        # Agar grade None bo'lsa, barcha testlarni ko'rsatish yoki bo'sh ro'yxat qaytarish
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.warning(f'Student {request.user.username} has no grade set')
                        test_ids = []
                    else:
                        with connection.cursor() as cursor:
                            cursor.execute(
                                "SELECT id FROM tests_app_test WHERE is_active = 1 AND grade = %s ORDER BY created_at DESC",
                                [student_grade]
                            )
                            test_ids = [row[0] for row in cursor.fetchall()]
                except Exception as db_error:
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.error(f'Error fetching test IDs (student): {str(db_error)}')
                    test_ids = []
                
                # Test obyektlarini xavfsiz yuklash
                tests = []
                for test_id in test_ids:
                    try:
                        test = Test.objects.select_related('created_by').get(id=test_id)
                        tests.append(test)
                    except Exception:
                        continue
                
                test_data = []
                for test in tests:
                    try:
                        attempt = TestAttempt.objects.filter(test=test, student=request.user).first()
                        test_data.append({
                            'id': test.id,
                            'title': test.title,
                            'subject': test.subject,
                            'description': test.description or '',
                            'grade': test.grade,
                            'time_limit': test.time_limit,
                            'max_attempts': test.max_attempts,
                            'total_questions': test.total_questions,
                            'has_attempted': attempt is not None,
                            'attempt_score': round(attempt.percentage, 1) if attempt and attempt.is_completed else None,
                            'can_attempt': (attempt is None or not attempt.is_completed) and test.is_active,
                            'created_by': test.created_by.get_full_name() if test.created_by and hasattr(test.created_by, 'get_full_name') else (test.created_by.username if test.created_by else 'Noma\'lum'),
                            'created_at': test.created_at.isoformat() if test.created_at else '',
                            'start_time': test.start_time.isoformat() if hasattr(test, 'start_time') and test.start_time else None,
                            'end_time': test.end_time.isoformat() if hasattr(test, 'end_time') and test.end_time else None,
                        })
                    except Exception as e:
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.error(f'Error processing test {test.id}: {str(e)}')
                        continue
                
                import logging
                logger = logging.getLogger(__name__)
                logger.info(f'Student test_list_view: Found {len(test_data)} tests for grade {student_grade}')
                
                return JsonResponse({
                    'tests': test_data,
                    'user_role': 'student'
                })
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f'Error in test_list_view (student): {str(e)}', exc_info=True)
                return JsonResponse({
                    'error': 'Testlarni yuklashda xatolik yuz berdi',
                    'detail': str(e)
                }, status=500)
            
        elif request.user.role == 'teacher':
            try:
                # Raw SQL yordamida test ID'larni olish
                test_ids = []
                try:
                    with connection.cursor() as cursor:
                        cursor.execute(
                            "SELECT id FROM tests_app_test WHERE created_by_id = %s ORDER BY created_at DESC",
                            [request.user.id]
                        )
                        test_ids = [row[0] for row in cursor.fetchall()]
                except Exception as db_error:
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.error(f'Error fetching test IDs (teacher): {str(db_error)}')
                    test_ids = []
                
                # Test obyektlarini xavfsiz yuklash
                tests = []
                for test_id in test_ids:
                    try:
                        test = Test.objects.select_related('created_by').get(id=test_id)
                        tests.append(test)
                    except Exception:
                        continue
                
                test_data = []
                for test in tests:
                    try:
                        attempt_count = TestAttempt.objects.filter(test=test, is_completed=True).count()
                        test_data.append({
                            'id': test.id,
                            'title': test.title,
                            'subject': test.subject,
                            'description': test.description or '',
                            'grade': test.grade,
                            'total_questions': test.total_questions,
                            'is_active': test.is_active,
                            'created_at': test.created_at.isoformat() if test.created_at else '',
                            'created_by': test.created_by.get_full_name() if test.created_by and hasattr(test.created_by, 'get_full_name') else (test.created_by.username if test.created_by else 'Noma\'lum'),
                            'attempt_count': attempt_count,
                            'max_attempts': test.max_attempts,
                            'time_limit': test.time_limit,
                        })
                    except Exception as e:
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.error(f'Error processing test {test.id}: {str(e)}')
                        continue
                
                return JsonResponse({
                    'tests': test_data,
                    'user_role': 'teacher'
                })
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f'Error in test_list_view (teacher): {str(e)}', exc_info=True)
                return JsonResponse({
                    'error': 'Testlarni yuklashda xatolik yuz berdi',
                    'detail': str(e)
                }, status=500)
        
        elif request.user.role == 'admin':
            # Admin barcha testlarni ko'radi
            try:
                # Raw SQL yordamida test ID'larni olish
                test_ids = []
                try:
                    with connection.cursor() as cursor:
                        cursor.execute("SELECT id FROM tests_app_test ORDER BY created_at DESC")
                        test_ids = [row[0] for row in cursor.fetchall()]
                except Exception as db_error:
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.error(f'Error fetching test IDs (admin): {str(db_error)}')
                    test_ids = []
                
                # Test obyektlarini xavfsiz yuklash
                tests = []
                for test_id in test_ids:
                    try:
                        test = Test.objects.select_related('created_by').get(id=test_id)
                        tests.append(test)
                    except Exception:
                        continue
                
                test_data = []
                for test in tests:
                    try:
                        attempt_count = TestAttempt.objects.filter(test=test, is_completed=True).count()
                        test_data.append({
                            'id': test.id,
                            'title': test.title,
                            'subject': test.subject,
                            'description': test.description or '',
                            'grade': test.grade,
                            'total_questions': test.total_questions,
                            'is_active': test.is_active,
                            'created_at': test.created_at.isoformat() if test.created_at else '',
                            'created_by': test.created_by.get_full_name() if test.created_by and hasattr(test.created_by, 'get_full_name') else (test.created_by.username if test.created_by else 'Noma\'lum'),
                            'attempt_count': attempt_count,
                            'max_attempts': test.max_attempts,
                            'time_limit': test.time_limit,
                        })
                    except Exception as e:
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.error(f'Error processing test {test.id}: {str(e)}')
                        continue
                
                import logging
                logger = logging.getLogger(__name__)
                logger.info(f'Admin test_list_view: Found {len(test_data)} tests')
                
                return JsonResponse({
                    'tests': test_data,
                    'user_role': 'admin'
                })
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f'Error in test_list_view (admin): {str(e)}', exc_info=True)
                return JsonResponse({
                    'error': 'Testlarni yuklashda xatolik yuz berdi',
                    'detail': str(e)
                }, status=500)
    
    return render(request, 'tests_app/test_list.html')

@login_required
@require_http_methods(["POST"])
def create_test(request):
    """Create a new test - Teachers only"""
    if request.user.role != 'teacher':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    try:
        # FormData yoki JSON formatini qo'llab-quvvatlash
        if request.content_type and 'multipart/form-data' in request.content_type:
            # FormData formatida
            title = request.POST.get('title')
            subject = request.POST.get('subject')
            grade = request.POST.get('grade')
            time_limit = request.POST.get('time_limit')
            
            if not all([title, subject, grade, time_limit]):
                return JsonResponse({'error': 'Barcha majburiy maydonlar to\'ldirilishi kerak'}, status=400)
            
            with transaction.atomic():
                test = Test.objects.create(
                    title=title,
                    description=request.POST.get('description', ''),
                    subject=subject,
                    grade=int(grade),
                    time_limit=int(time_limit),
                    created_by=request.user,
                    max_attempts=int(request.POST.get('max_attempts', 1)),
                    show_results=request.POST.get('show_results', 'true') == 'true',
                    shuffle_questions=request.POST.get('shuffle_questions', 'false') == 'true'
                )
                
                # Savollarni qayta ishlash
                question_texts = request.POST.getlist('question_text[]')
                question_types = request.POST.getlist('question_type[]')
                points_list = request.POST.getlist('points[]')
                explanations = request.POST.getlist('explanation[]')
                question_images = request.FILES.getlist('question_image[]')
                
                for i in range(len(question_texts)):
                    question = Question.objects.create(
                        test=test,
                        question_text=question_texts[i],
                        question_type=question_types[i],
                        points=float(points_list[i]) if points_list[i] else 1.0,
                        order=i + 1,
                        explanation=explanations[i] if i < len(explanations) else ''
                    )
                    
                    # Rasm yuklash
                    if i < len(question_images) and question_images[i]:
                        question.image = question_images[i]
                        question.save()
                    
                    # Variantlarni qayta ishlash
                    if question_types[i] in ['single_choice', 'multiple_choice']:
                        choices = request.POST.getlist(f'choices_{i+1}[]')
                        correct_choice = request.POST.get(f'correct_choice_{i+1}')
                        
                        for j, choice_text in enumerate(choices):
                            if choice_text.strip():
                                is_correct = str(j) == correct_choice
                                Choice.objects.create(
                                    question=question,
                                    choice_text=choice_text,
                                    is_correct=is_correct
                                )
            
            return JsonResponse({
                'success': True,
                'message': 'Test muvaffaqiyatli yaratildi',
                'test_id': test.id
            })
        else:
            # JSON formatida (eski format)
            data = json.loads(request.body)
            
            required_fields = ['title', 'subject', 'grade', 'time_limit']
            for field in required_fields:
                if not data.get(field):
                    return JsonResponse({'error': f'{field} is required'}, status=400)
            
            with transaction.atomic():
                test = Test.objects.create(
                    title=data['title'],
                    description=data.get('description', ''),
                    subject=data['subject'],
                    grade=int(data['grade']),
                    time_limit=int(data['time_limit']),
                    created_by=request.user,
                    start_time=data.get('start_time'),
                    end_time=data.get('end_time'),
                    max_attempts=data.get('max_attempts', 1),
                    show_results=data.get('show_results', True),
                    shuffle_questions=data.get('shuffle_questions', False)
                )
                
                questions_data = data.get('questions', [])
                for i, q_data in enumerate(questions_data):
                    question = Question.objects.create(
                        test=test,
                        question_text=q_data['question_text'],
                        question_type=q_data['question_type'],
                        points=float(q_data.get('points', 1.0)),
                        order=i + 1,
                        explanation=q_data.get('explanation', '')
                    )
                    
                    # Base64 rasm yuklash (JSON formatida)
                    if q_data.get('image_base64'):
                        import base64
                        from django.core.files.base import ContentFile
                        from django.core.files.uploadedfile import InMemoryUploadedFile
                        import io
                        
                        try:
                            format, imgstr = q_data['image_base64'].split(';base64,')
                            ext = format.split('/')[-1]
                            data_img = ContentFile(base64.b64decode(imgstr), name=f'question_{question.id}.{ext}')
                            question.image = data_img
                            question.save()
                        except Exception as e:
                            pass  # Rasm yuklashda xatolik bo'lsa, davom et
                    
                    if q_data['question_type'] in ['single_choice', 'multiple_choice']:
                        choices_data = q_data.get('choices', [])
                        for choice_data in choices_data:
                            Choice.objects.create(
                                question=question,
                                choice_text=choice_data['text'],
                                is_correct=choice_data.get('is_correct', False)
                            )
            
            return JsonResponse({
                'success': True,
                'message': 'Test created successfully',
                'test_id': test.id
            })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Error creating test: {str(e)}', exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def take_test_view(request, test_id):
    """Start or continue taking a test - Students only"""
    if request.user.role != 'student':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    test = get_object_or_404(Test, id=test_id, is_active=True)
    
    if test.grade != request.user.grade:
        return JsonResponse({'error': 'This test is not for your grade'}, status=403)
    
    now = timezone.now()
    if test.start_time and now < test.start_time:
        return JsonResponse({'error': 'Test has not started yet'}, status=403)
    
    if test.end_time and now > test.end_time:
        return JsonResponse({'error': 'Test has ended'}, status=403)
    
    # Migration qo'llanmagan bo'lsa ham ishlashi uchun
    is_paused = getattr(test, 'is_paused', False)
    if is_paused:
        return JsonResponse({'error': 'Test hozir pauza qilingan. Iltimos, kuting...'}, status=403)
    
    if request.method == 'POST':
        existing_attempt = TestAttempt.objects.filter(test=test, student=request.user).first()
        if existing_attempt and existing_attempt.is_completed:
            return JsonResponse({'error': 'You have already completed this test'}, status=400)
        
        if not existing_attempt:
            attempt = TestAttempt.objects.create(test=test, student=request.user)
        else:
            attempt = existing_attempt
        
        # Barcha savollarni olish
        all_questions = list(test.questions.all())
        
        # Agar 50 dan ko'p savol bo'lsa, random 50 tasini tanlab olish
        MAX_QUESTIONS = 50
        if len(all_questions) > MAX_QUESTIONS:
            # Random 50 ta savolni tanlab olish (har safar turli savollar)
            selected_questions = random.sample(all_questions, MAX_QUESTIONS)
        else:
            selected_questions = all_questions
        
        # Savollarni aralashtirish (agar shuffle_questions True bo'lsa)
        if test.shuffle_questions:
            random.shuffle(selected_questions)
        
        questions_data = []
        for question in selected_questions:
            q_data = {
                'id': question.id,
                'question_text': question.question_text,
                'question_type': question.question_type,
                'points': question.points,
                'image_url': question.image.url if question.image else None
            }
            
            if question.question_type in ['single_choice', 'multiple_choice']:
                # Variantlarni olish va aralashtirish (har safar turli tartibda)
                choices = list(question.choices.all())
                random.shuffle(choices)  # Variantlarni aralashtirish
                
                q_data['choices'] = [{
                    'id': choice.id,
                    'text': choice.choice_text
                } for choice in choices]
            
            questions_data.append(q_data)
        
        return JsonResponse({
            'attempt_id': attempt.id,
            'questions': questions_data,
            'time_limit': test.time_limit,
            'started_at': attempt.started_at.isoformat(),
            'server_time': timezone.now().isoformat()
        })
    
    # GET request uchun server vaqtini qaytarish
    if request.method == 'GET' and request.headers.get('Accept') == 'application/json':
        return JsonResponse({
            'server_time': timezone.now().isoformat()
        })
    
    return render(request, 'tests_app/take_test.html', {'test': test})

@login_required
@require_http_methods(["POST"])
def submit_answer(request, attempt_id):
    """Submit answer for a question"""
    if request.user.role != 'student':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    try:
        data = json.loads(request.body)
        attempt = get_object_or_404(TestAttempt, id=attempt_id, student=request.user)
        
        if attempt.is_completed:
            return JsonResponse({'error': 'Test already completed'}, status=400)
        
        question_id = data.get('question_id')
        question = get_object_or_404(Question, id=question_id, test=attempt.test)
        
        answer, created = Answer.objects.get_or_create(
            attempt=attempt,
            question=question
        )
        
        answer.selected_choices.clear()
        answer.text_answer = ''
        
        if question.question_type == 'text_answer':
            answer.text_answer = data.get('text_answer', '')
        else:
            choice_ids = data.get('choice_ids', [])
            if choice_ids:
                choices = Choice.objects.filter(id__in=choice_ids, question=question)
                answer.selected_choices.set(choices)
        
        answer.save()
        
        return JsonResponse({'message': 'Answer saved'})
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@require_http_methods(["POST"])
def finish_test(request, attempt_id):
    """Finish the test and calculate score"""
    if request.user.role != 'student':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    try:
        attempt = get_object_or_404(TestAttempt, id=attempt_id, student=request.user)
        
        if attempt.is_completed:
            return JsonResponse({'error': 'Test already completed'}, status=400)
        
        attempt.finished_at = timezone.now()
        attempt.is_completed = True
        attempt.time_taken = attempt.finished_at - attempt.started_at
        
        results = attempt.calculate_score()
        
        correct_answers = 0
        incorrect_answers = 0
        unanswered = 0
        
        # Faqat tanlangan savollarni tekshirish (Answer modelida saqlangan savollar)
        answered_questions = Answer.objects.filter(attempt=attempt).select_related('question')
        answered_question_ids = set(answer.question_id for answer in answered_questions)
        
        # Barcha test savollarini olish
        all_test_questions_count = attempt.test.questions.count()
        
        # Agar 50 dan ko'p savol bo'lsa, faqat javob berilgan savollarni tekshirish
        # Aks holda barcha savollarni tekshirish
        if all_test_questions_count > 50:
            # Faqat javob berilgan savollarni tekshirish
            for answer in answered_questions:
                if answer.is_correct():
                    correct_answers += 1
                else:
                    incorrect_answers += 1
            
            # Javob berilmagan savollar soni (faqat tanlangan 50 ta savoldan)
            unanswered = 50 - len(answered_question_ids)
        else:
            # Barcha savollarni tekshirish
            for question in attempt.test.questions.all():
                answer = Answer.objects.filter(attempt=attempt, question=question).first()
                if answer:
                    if answer.is_correct():
                        correct_answers += 1
                    else:
                        incorrect_answers += 1
                else:
                    unanswered += 1
        
        test_result = TestResult.objects.create(
            attempt=attempt,
            correct_answers=correct_answers,
            incorrect_answers=incorrect_answers,
            unanswered=unanswered
        )
        test_result.grade = test_result.calculate_grade()
        test_result.save()
        
        attempt.save()
        
        completion_message = "Test yakunlandi!"
        if results.get('all_answered', False):
            completion_message = f"Ajoyib! Barcha {results['total_questions']} ta savolga javob berdingiz!"
        else:
            completion_message = f"Test yakunlandi. {results['answered_count']}/{results['total_questions']} ta savolga javob berildi."
        
        return JsonResponse({
            'message': completion_message,
            'results': {
                'score': results['score'],
                'total_points': results['total_points'],
                'percentage': results['percentage'],
                'grade': test_result.grade,
                'correct_answers': correct_answers,
                'incorrect_answers': incorrect_answers,
                'unanswered': unanswered,
                'time_taken': str(attempt.time_taken),
                'all_answered': results.get('all_answered', False),
                'answered_count': results.get('answered_count', 0),
                'total_questions': results.get('total_questions', 0),
                'incorrect_questions': results.get('incorrect_questions', [])
            }
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def test_results_view(request, test_id):
    """View test results - Teachers can see all, students see their own"""
    test = get_object_or_404(Test, id=test_id)
    
    if request.headers.get('Accept') == 'application/json':
        if request.user.role == 'student':
            if test.grade != request.user.grade:
                return JsonResponse({'error': 'Access denied'}, status=403)
            
            attempt = TestAttempt.objects.filter(test=test, student=request.user).first()
            if not attempt or not attempt.is_completed:
                return JsonResponse({'error': 'Test not completed'}, status=404)
            
            results = attempt.calculate_score() if hasattr(attempt, 'calculate_score') else {}
            correct_answers = attempt.result.correct_answers if hasattr(attempt, 'result') else 0
            incorrect_answers = attempt.result.incorrect_answers if hasattr(attempt, 'result') else 0
            unanswered = attempt.result.unanswered if hasattr(attempt, 'result') else 0
            result_data = {
                'student': request.user.username,
                'score': attempt.score,
                'total_points': attempt.total_points,
                'percentage': attempt.percentage,
                'grade': attempt.result.grade if hasattr(attempt, 'result') else '',
                'time_taken': str(attempt.time_taken),
                'finished_at': attempt.finished_at.isoformat(),
                'correct_answers': correct_answers,
                'incorrect_answers': incorrect_answers,
                'unanswered': unanswered,
                'all_answered': results.get('all_answered', False),
                'answered_count': results.get('answered_count', 0),
                'total_questions': results.get('total_questions', 0),
                'incorrect_questions': results.get('incorrect_questions', [])
            }
            return JsonResponse({'result': result_data})
        
        elif request.user.role == 'teacher' and test.created_by == request.user:
            attempts = TestAttempt.objects.filter(test=test, is_completed=True).select_related('student', 'result').order_by('student__grade', 'student__class_name', 'student__first_name', 'student__last_name')
            
            results_data = []
            for attempt in attempts:
                results_data.append({
                    'student': {
                        'username': attempt.student.username,
                        'first_name': attempt.student.first_name,
                        'last_name': attempt.student.last_name,
                        'student_id': attempt.student.student_id,
                        'class_name': attempt.student.class_name,
                        'grade': attempt.student.grade
                    },
                    'score': attempt.score,
                    'total_points': attempt.total_points,
                    'percentage': attempt.percentage,
                    'grade': attempt.result.grade if hasattr(attempt, 'result') else '',
                    'time_taken': str(attempt.time_taken),
                    'finished_at': attempt.finished_at.isoformat(),
                    'correct_answers': attempt.result.correct_answers if hasattr(attempt, 'result') else 0,
                    'incorrect_answers': attempt.result.incorrect_answers if hasattr(attempt, 'result') else 0
                })
            
            return JsonResponse({'results': results_data})
        
        else:
            return JsonResponse({'error': 'Access denied'}, status=403)
    
    return render(request, 'tests_app/test_results.html', {
        'test': test,
        'user_role': request.user.role
    })

@login_required
def export_results(request, test_id):
    """Export test results to Excel - Teachers only"""
    if request.user.role != 'teacher':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    test = get_object_or_404(Test, id=test_id, created_by=request.user)
    attempts = TestAttempt.objects.filter(test=test, is_completed=True).select_related('student', 'result').order_by('student__grade', 'student__class_name', 'student__first_name', 'student__last_name')
    
    
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Results"
    
    # Header qo'shish
    headers = [
        'Student Username', 'First Name', 'Last Name', 'Student ID', 'Grade', 
        'Class', 'Score', 'Total Points', 'Percentage', 'Grade Result',
        'Correct Answers', 'Incorrect Answers', 'Unanswered', 'Time Taken', 'Finished At'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Ma'lumotlarni qo'shish
    for row, attempt in enumerate(attempts, 2):
        data = [
            attempt.student.username,
            attempt.student.first_name,
            attempt.student.last_name,
            attempt.student.student_id or '',
            attempt.student.grade or '',
            attempt.student.class_name or '',
            attempt.score,
            attempt.total_points,
            attempt.percentage,
            attempt.result.grade if hasattr(attempt, 'result') else '',
            attempt.result.correct_answers if hasattr(attempt, 'result') else 0,
            attempt.result.incorrect_answers if hasattr(attempt, 'result') else 0,
            attempt.result.unanswered if hasattr(attempt, 'result') else 0,
            str(attempt.time_taken),
            attempt.finished_at.strftime('%Y-%m-%d %H:%M:%S')
        ]
        
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Excel faylini qaytarish
    from django.http import HttpResponse
    from io import BytesIO
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{test.title}_results.xlsx"'
    
    return response

@login_required
def upload_questions(request, test_id):
    """Upload questions from Excel file - Teachers only"""
    if request.user.role != 'teacher':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    test = get_object_or_404(Test, id=test_id, created_by=request.user)
    
    if request.method == 'POST':
        try:
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                return JsonResponse({'error': 'No file uploaded'}, status=400)
            
            # Excel faylini openpyxl bilan o'qish
            from openpyxl import load_workbook
            
            wb = load_workbook(excel_file)
            ws = wb.active
            
            # Header qatorini o'qish
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(cell.value.lower().replace(' ', '_'))
            
            # Kerakli ustunlarni tekshirish
            required_columns = ['question_text', 'question_type', 'points']
            for col in required_columns:
                if col not in headers:
                    return JsonResponse({'error': f'Missing column: {col}'}, status=400)
            
            questions_created = 0
            
            with transaction.atomic():
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
                    if not row[0]:  # question_text bo'sh bo'lsa
                        continue
                    
                    row_data = dict(zip(headers, row))
                    
                    question = Question.objects.create(
                        test=test,
                        question_text=row_data['question_text'],
                        question_type=row_data['question_type'],
                        points=float(row_data.get('points', 1.0)),
                        order=row_num,
                        explanation=row_data.get('explanation', '')
                    )
                    
                    # Javob variantlarini qo'shish
                    if row_data['question_type'] in ['single_choice', 'multiple_choice']:
                        for i in range(1, 6):  # 5 tagacha variant
                            choice_key = f'choice_{i}'
                            correct_key = f'choice_{i}_correct'
                            
                            if choice_key in row_data and row_data[choice_key]:
                                is_correct = bool(row_data.get(correct_key, False))
                                Choice.objects.create(
                                    question=question,
                                    choice_text=row_data[choice_key],
                                    is_correct=is_correct
                                )
                    
                    questions_created += 1
            
            return JsonResponse({'message': f'{questions_created} questions uploaded successfully'})
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return render(request, 'tests_app/upload_questions.html', {'test': test})

@login_required
def create_test_view(request):
    """Create new test - for teachers and admins"""
    if request.user.role not in ['teacher', 'admin']:
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    if request.method == 'GET':
        return render(request, 'tests_app/create_test.html')
    
    if request.method == 'POST':
        try:
            # Majburiy maydonlarni tekshirish
            title = request.POST.get('title', '').strip()
            subject = request.POST.get('subject', '').strip()
            grade = request.POST.get('grade', '').strip()
            time_limit = request.POST.get('time_limit', '45').strip()
            
            if not title or not subject or not grade:
                return JsonResponse({
                    'success': False, 
                    'error': 'Test nomi, fan va sinf majburiy maydonlar!'
                }, status=400)
            
            # Savollar mavjudligini tekshirish
            question_texts = request.POST.getlist('question_text[]')
            if not question_texts or not any(q.strip() for q in question_texts):
                return JsonResponse({
                    'success': False,
                    'error': 'Kamida bitta savol qo\'shilishi kerak!'
                }, status=400)
            
            with transaction.atomic():
                # Test yaratish - is_paused maydonini xavfsiz qo'shish
                # Asosiy maydonlar
                test_kwargs = {
                    'title': title,
                    'description': request.POST.get('description', '').strip(),
                    'subject': subject,
                    'grade': int(grade),
                    'time_limit': int(time_limit),
                    'max_attempts': int(request.POST.get('max_attempts', 1)),
                    'show_results': request.POST.get('show_results') == 'on' or request.POST.get('show_results') == 'true',
                    'is_active': request.POST.get('is_active') == 'on' or request.POST.get('is_active') == 'true',
                    'shuffle_questions': request.POST.get('shuffle_questions') == 'on' or request.POST.get('shuffle_questions') == 'true',
                    'created_by': request.user
                }
                
                # is_paused va paused_at maydonlarini xavfsiz qo'shish
                # Agar maydonlar mavjud bo'lmasa, ularni qo'shmasdan yaratish
                try:
                    # Avval model maydonlarini tekshirish
                    from django.db import models
                    test_fields = [f.name for f in Test._meta.get_fields() if isinstance(f, models.Field)]
                    if 'is_paused' in test_fields:
                        test_kwargs['is_paused'] = False
                    if 'paused_at' in test_fields:
                        test_kwargs['paused_at'] = None
                except:
                    # Agar maydonlarni tekshirib bo'lmasa, ularni qo'shmasdan davom etish
                    pass
                
                # Test yaratish
                try:
                    test = Test.objects.create(**test_kwargs)
                except Exception as db_error:
                    # Agar is_paused maydoni muammosi bo'lsa, uni qo'shmasdan qayta urinib ko'rish
                    error_str = str(db_error).lower()
                    if 'is_paused' in error_str or 'paused_at' in error_str or 'no such column' in error_str:
                        # is_paused va paused_at maydonlarini olib tashlash
                        test_kwargs.pop('is_paused', None)
                        test_kwargs.pop('paused_at', None)
                        test = Test.objects.create(**test_kwargs)
                    else:
                        # Boshqa xatolik bo'lsa, qayta tashlash
                        raise
                
                # Savollar qo'shish
                question_texts = request.POST.getlist('question_text[]')
                question_types = request.POST.getlist('question_type[]')
                points_list = request.POST.getlist('points[]')
                explanations = request.POST.getlist('explanation[]')
                question_images = request.FILES.getlist('question_image[]')
                
                question_order = 0
                for i, question_text in enumerate(question_texts):
                    if not question_text.strip():
                        continue
                    
                    question_order += 1
                    question = Question.objects.create(
                        test=test,
                        question_text=question_text.strip(),
                        question_type=question_types[i] if i < len(question_types) else 'single_choice',
                        points=float(points_list[i]) if i < len(points_list) and points_list[i] else 1.0,
                        order=question_order,
                        explanation=explanations[i].strip() if i < len(explanations) and explanations[i] else ''
                    )
                    
                    # Rasm yuklash
                    if i < len(question_images) and question_images[i]:
                        question.image = question_images[i]
                        question.save()
                    
                    # Javob variantlarini qo'shish
                    question_type = question_types[i] if i < len(question_types) else 'single_choice'
                    if question_type != 'text_answer':
                        choices_key = f'choices_{i+1}[]'
                        correct_key = f'correct_choice_{i+1}'
                        
                        choices = request.POST.getlist(choices_key)
                        correct_index = request.POST.get(correct_key, '')
                        
                        # Kamida 2 ta variant bo'lishi kerak
                        valid_choices = [c.strip() for c in choices if c.strip()]
                        if len(valid_choices) < 2:
                            raise ValueError(f'Savol {question_order} uchun kamida 2 ta javob varianti bo\'lishi kerak!')
                        
                        # To'g'ri javobni tekshirish
                        if question_type == 'single_choice':
                            if not correct_index or correct_index == '':
                                raise ValueError(f'Savol {question_order} uchun to\'g\'ri javob tanlanishi kerak!')
                        
                        for j, choice_text in enumerate(choices):
                            if choice_text.strip():
                                if question_type == 'single_choice':
                                    is_correct = str(j) == correct_index
                                else:  # multiple_choice
                                    # Multiple choice uchun checkbox'lar ishlatiladi
                                    checkbox_name = f'correct_choice_{i+1}_{j}'
                                    is_correct = request.POST.get(checkbox_name) == 'on'
                                
                                Choice.objects.create(
                                    question=question,
                                    choice_text=choice_text.strip(),
                                    is_correct=is_correct
                                )
                        
                        # Single choice uchun kamida bitta to'g'ri javob bo'lishi kerak
                        if question_type == 'single_choice':
                            has_correct = Choice.objects.filter(question=question, is_correct=True).exists()
                            if not has_correct:
                                raise ValueError(f'Savol {question_order} uchun kamida bitta to\'g\'ri javob bo\'lishi kerak!')
                
                return JsonResponse({
                    'success': True, 
                    'message': 'Test muvaffaqiyatli yaratildi!',
                    'test_id': test.id
                })
                
        except ValueError as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f'Error creating test: {str(e)}', exc_info=True)
            return JsonResponse({'success': False, 'error': f'Xatolik: {str(e)}'}, status=500)

@login_required
def test_info_view(request, test_id):
    """Get test information for display purposes"""
    test = get_object_or_404(Test, id=test_id)
    
    # Check access permissions
    if request.user.role == 'student' and test.grade != request.user.grade:
        return JsonResponse({'error': 'Access denied'}, status=403)
    elif request.user.role == 'teacher' and test.created_by != request.user:
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    return JsonResponse({
        'title': test.title,
        'description': test.description,
        'subject': test.subject,
        'grade': test.grade,
        'time_limit': test.time_limit,
        'max_attempts': test.max_attempts,
        'total_questions': test.total_questions,
        'created_by': test.created_by.get_full_name() or test.created_by.username,
        'created_at': test.created_at.isoformat(),
        'start_time': test.start_time.isoformat() if test.start_time else None,
        'is_paused': getattr(test, 'is_paused', False),
        'paused_at': test.paused_at.isoformat() if hasattr(test, 'paused_at') and test.paused_at else None,
        'end_time': test.end_time.isoformat() if test.end_time else None,
    })

@login_required
def all_results_view(request):
    """Barcha test natijalarini ko'rsatish - Admin va Teacher uchun"""
    if request.user.role not in ['admin', 'teacher']:
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    # Excel export uchun
    if request.method == 'GET' and request.GET.get('export') == 'excel':
        grade_filter = request.GET.get('grade', None)
        
        # Admin barcha natijalarni ko'radi, Teacher faqat o'z testlari natijalarini
        if request.user.role == 'admin':
            attempts = TestAttempt.objects.filter(is_completed=True).select_related(
                'student', 'test', 'result'
            )
        else:  # teacher
            attempts = TestAttempt.objects.filter(
                test__created_by=request.user, 
                is_completed=True
            ).select_related('student', 'test', 'result')
        
        # Sinf bo'yicha filter
        if grade_filter:
            attempts = attempts.filter(student__grade=grade_filter)
        
        attempts = attempts.order_by('student__grade', 'student__class_name', 'student__first_name', '-finished_at')
        
        # Excel fayl yaratish
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Natijalari"
        
        # Header qo'shish
        headers = [
            'O\'quvchi FIO', 'Username', 'Student ID', 'Sinf', 'Sinif', 
            'Test Nomi', 'Fan', 'Ball', 'Umumiy Ball', 'Foiz', 'Baho',
            'To\'g\'ri Javoblar', 'Noto\'g\'ri Javoblar', 'Javobsiz', 
            'Vaqt', 'Sana va Vaqt'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, size=12, color="FFFFFF")
        
        # Ma'lumotlarni qo'shish
        for row, attempt in enumerate(attempts, 2):
            percentage = attempt.percentage or 0
            if percentage >= 81:
                grade = "A'lo"
            elif percentage >= 61:
                grade = "Yaxshi"
            elif percentage >= 31:
                grade = "Qoniqarli"
            else:
                grade = "Qoniqarsiz"
            
            student_name = f"{attempt.student.first_name} {attempt.student.last_name}"
            
            data = [
                student_name,
                attempt.student.username,
                attempt.student.student_id or '',
                attempt.student.grade or '',
                attempt.student.class_name or '',
                attempt.test.title,
                attempt.test.subject,
                attempt.score,
                attempt.total_points,
                f"{percentage:.1f}%",
                grade,
                attempt.result.correct_answers if hasattr(attempt, 'result') else 0,
                attempt.result.incorrect_answers if hasattr(attempt, 'result') else 0,
                attempt.result.unanswered if hasattr(attempt, 'result') else 0,
                str(attempt.time_taken),
                attempt.finished_at.strftime('%Y-%m-%d %H:%M:%S')
            ]
            
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Column width'ni sozlash
        column_widths = [25, 15, 12, 8, 10, 30, 15, 8, 12, 10, 12, 12, 12, 10, 15, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        
        # Excel faylini qaytarish
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = 'barcha_test_natijalari'
        if grade_filter:
            filename += f'_sinf_{grade_filter}'
        filename += '.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
    
    if request.method == 'GET' and request.headers.get('Accept') == 'application/json':
        # Admin barcha natijalarni ko'radi, Teacher faqat o'z testlari natijalarini
        if request.user.role == 'admin':
            attempts = TestAttempt.objects.filter(is_completed=True).select_related(
                'student', 'test', 'result'
            ).order_by('student__grade', 'student__class_name', 'student__first_name', '-finished_at')
        else:  # teacher
            attempts = TestAttempt.objects.filter(
                test__created_by=request.user, 
                is_completed=True
            ).select_related('student', 'test', 'result').order_by(
                'student__grade', 'student__class_name', 'student__first_name', '-finished_at'
            )
        
        results_data = []
        for attempt in attempts:
            # Calculate grade based on percentage
            percentage = attempt.percentage or 0
            if percentage >= 81:
                grade = "A'lo"
            elif percentage >= 61:
                grade = "Yaxshi"
            elif percentage >= 31:
                grade = "Qoniqarli"
            else:
                grade = "Qoniqarsiz"
            results_data.append({
                'test': {
                    'id': attempt.test.id,
                    'title': attempt.test.title,
                    'subject': attempt.test.subject,
                    'grade': attempt.test.grade,
                    'created_by': attempt.test.created_by.get_full_name() or attempt.test.created_by.username
                },
                'student': {
                    'id': attempt.student.id,
                    'username': attempt.student.username,
                    'first_name': attempt.student.first_name,
                    'last_name': attempt.student.last_name,
                    'student_id': attempt.student.student_id,
                    'class_name': attempt.student.class_name,
                    'grade': attempt.student.grade
                },
                'score': attempt.score,
                'total_points': attempt.total_points,
                'percentage': attempt.percentage,
                'grade': grade,
                'time_taken': str(attempt.time_taken),
                'finished_at': attempt.finished_at.isoformat(),
                'correct_answers': attempt.result.correct_answers if hasattr(attempt, 'result') else 0,
                'incorrect_answers': attempt.result.incorrect_answers if hasattr(attempt, 'result') else 0,
                'unanswered': attempt.result.unanswered if hasattr(attempt, 'result') else 0
            })
        
        return JsonResponse({'results': results_data})
    
    return render(request, 'tests_app/all_results.html', {
        'user_role': request.user.role
    })

@login_required
@require_http_methods(["POST"])
def request_retake_view(request, test_id):
    """O'quvchi qayta ishlash so'rovi yuborish"""
    if request.user.role != 'student':
        return JsonResponse({'error': 'Faqat o\'quvchilar qayta ishlash so\'rovi yuborishi mumkin'}, status=403)
    
    test = get_object_or_404(Test, id=test_id)
    
    try:
        # O'quvchining oxirgi attempt'ini topamiz
        attempt = TestAttempt.objects.filter(test=test, student=request.user, is_completed=True).last()
        if not attempt:
            return JsonResponse({'error': 'Siz hali bu testni topshirmadingiz'}, status=400)
        
        # Qayta ishlash so'rashi mumkinmi?
        if not attempt.can_request_retake():
            return JsonResponse({'error': 'Allaqachon qayta ishlash so\'rovi yuborilgan yoki kutilmoqda'}, status=400)
        
        data = json.loads(request.body)
        reason = data.get('reason', '').strip()
        
        if not reason:
            return JsonResponse({'error': 'Qayta ishlash sababi kiritilishi shart'}, status=400)
        
        # Qayta ishlash so'rovini yaratamiz
        retake_request = TestRetakeRequest.objects.create(
            student=request.user,
            test=test,
            previous_attempt=attempt,
            reason=reason
        )
        
        return JsonResponse({
            'message': 'Qayta ishlash so\'rovi muvaffaqiyatli yuborildi!',
            'request_id': retake_request.id
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Noto\'g\'ri JSON ma\'lumot'}, status=400)
    except Exception as e:
        return JsonResponse({'error': 'Xatolik yuz berdi'}, status=500)

@login_required
def retake_requests_view(request):
    """Admin va O'qituvchilar qayta ishlash so'rovlarini ko'rish va boshqarish"""
    if request.user.role not in ['admin', 'teacher']:
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    if request.method == 'GET' and request.headers.get('Accept') == 'application/json':
        try:
            # JSON API so'rovi
            status_filter = request.GET.get('status', 'all')
            
            # Test modelini yuklashda xatolik bo'lmasligi uchun xavfsir so'rov
            # Raw SQL yoki values() ishlatish orqali Test modelini to'liq yuklamaslik
            try:
                # Avval TestRetakeRequest'larni olish, keyin test ma'lumotlarini alohida olish
                requests_qs = TestRetakeRequest.objects.select_related(
                    'student', 'previous_attempt', 'approved_by'
                ).order_by('-created_at')
                
                # O'qituvchi uchun faqat o'z testlari so'rovlari
                # test__created_by filtri Test modelini yuklaydi, shuning uchun xavfsiz qilamiz
                if request.user.role == 'teacher':
                    try:
                        # Raw SQL yordamida test ID'larni olish (Test modelini to'liq yuklamaslik)
                        from django.db import connection
                        with connection.cursor() as cursor:
                            cursor.execute(
                                "SELECT id FROM tests_app_test WHERE created_by_id = %s",
                                [request.user.id]
                            )
                            teacher_test_ids = [row[0] for row in cursor.fetchall()]
                        
                        if teacher_test_ids:
                            requests_qs = requests_qs.filter(test_id__in=teacher_test_ids)
                        else:
                            # O'qituvchining testlari yo'q
                            requests_qs = requests_qs.none()
                    except Exception as filter_error:
                        # Agar Test modelini yuklashda xatolik bo'lsa, bo'sh ro'yxat qaytarish
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.error(f'Error filtering teacher tests: {str(filter_error)}')
                        return JsonResponse({
                            'requests': [],
                            'total_count': 0
                        })
                
                if status_filter != 'all':
                    requests_qs = requests_qs.filter(status=status_filter)
            except Exception as db_error:
                # Database xatolik - migration qo'llanmagan bo'lishi mumkin
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f'Database error in retake_requests_view: {str(db_error)}', exc_info=True)
                # Bo'sh ro'yxat qaytarish
                return JsonResponse({
                    'requests': [],
                    'total_count': 0
                })
            
            requests_data = []
            for req in requests_qs:
                try:
                    # previous_attempt None bo'lishi mumkinligini tekshirish
                    previous_score = 0
                    previous_percentage = 0
                    if req.previous_attempt:
                        previous_score = req.previous_attempt.score or 0
                        previous_percentage = req.previous_attempt.percentage or 0
                    
                    # Test ma'lumotlarini xavfsiz olish
                    test_title = req.test.title if req.test else 'Noma\'lum test'
                    test_subject = req.test.subject if req.test else '-'
                    
                    requests_data.append({
                        'id': req.id,
                        'student_name': req.student.get_full_name() or req.student.username if req.student else 'Noma\'lum',
                        'student_username': req.student.username if req.student else '-',
                        'student_grade': req.student.grade or '-' if req.student else '-',
                        'student_class': req.student.class_name or '-' if req.student else '-',
                        'test_title': test_title,
                        'test_subject': test_subject,
                        'previous_score': previous_score,
                        'previous_percentage': previous_percentage,
                        'reason': req.reason or '',
                        'status': req.status,
                        'status_display': req.get_status_display(),
                        'admin_response': req.admin_response or '',
                        'approved_by': req.approved_by.get_full_name() if req.approved_by else None,
                        'created_at': req.created_at.isoformat() if req.created_at else '',
                        'updated_at': req.updated_at.isoformat() if req.updated_at else ''
                    })
                except Exception as e:
                    # Xatolikni log qilish va davom etish
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.error(f'Error processing retake request {req.id}: {str(e)}')
                    continue
            
            return JsonResponse({
                'requests': requests_data,
                'total_count': len(requests_data)
            })
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f'Error in retake_requests_view: {str(e)}', exc_info=True)
            return JsonResponse({
                'error': 'Ma\'lumotlarni yuklashda xatolik yuz berdi',
                'detail': str(e)
            }, status=500)
    
    # HTML template
    return render(request, 'tests_app/retake_requests.html', {
        'user_role': request.user.role
    })

@login_required
@require_http_methods(["POST"])
def handle_retake_request_view(request, request_id):
    """Admin va O'qituvchilar qayta ishlash so'rovini tasdiqlash yoki rad etish"""
    if request.user.role not in ['admin', 'teacher']:
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    retake_request = get_object_or_404(TestRetakeRequest, id=request_id)
    
    # O'qituvchi faqat o'z testlari so'rovlarini boshqarishi mumkin
    if request.user.role == 'teacher' and retake_request.test.created_by != request.user:
        return JsonResponse({'error': 'Siz faqat o\'z testlaringiz so\'rovlarini boshqarishingiz mumkin'}, status=403)
    
    if retake_request.status != 'pending':
        return JsonResponse({'error': 'Bu so\'rov allaqachon ko\'rib chiqilgan'}, status=400)
    
    try:
        data = json.loads(request.body)
        action = data.get('action')  # 'approve' yoki 'reject'
        admin_response = data.get('admin_response', '').strip()
        
        if action not in ['approve', 'reject']:
            return JsonResponse({'error': 'Noto\'g\'ri harakat'}, status=400)
        
        retake_request.status = 'approved' if action == 'approve' else 'rejected'
        retake_request.admin_response = admin_response
        retake_request.approved_by = request.user
        retake_request.save()
        
        return JsonResponse({
            'message': f'So\'rov {"tasdiqlandi" if action == "approve" else "rad etildi"}!',
            'status': retake_request.status
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Noto\'g\'ri JSON ma\'lumot'}, status=400)
    except Exception as e:
        return JsonResponse({'error': 'Xatolik yuz berdi'}, status=500)


@login_required
def open_test_for_student(request, test_id, student_id):
    """Admin tomonidan o'quvchi uchun testni qayta ochish"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Ruxsat berilmagan'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST so\'rov talab qilinadi'}, status=405)
    
    try:
        test = Test.objects.get(id=test_id)
        student = User.objects.get(id=student_id, role='student')
        
        # O'quvchining bu testdagi avvalgi urinishlarini tekshirish
        previous_attempts = TestAttempt.objects.filter(
            student=student,
            test=test
        ).count()
        
        # Yangi urinish yaratish (qayta ishlash imkoniyati)
        new_attempt = TestAttempt.objects.create(
            student=student,
            test=test,
            attempt_number=previous_attempts + 1,
            is_retake=True
        )
        
        # Agar qayta ishlash so'rovi mavjud bo'lsa, uni tasdiqlangan deb belgilash
        retake_request = TestRetakeRequest.objects.filter(
            student=student,
            test=test,
            status='approved'
        ).first()
        
        if retake_request:
            retake_request.is_used = True
            retake_request.save()
        
        return JsonResponse({
            'message': f'{student.get_full_name()} uchun "{test.title}" testi qayta ochildi!',
            'attempt_id': new_attempt.id,
            'attempt_number': new_attempt.attempt_number
        })
        
    except Test.DoesNotExist:
        return JsonResponse({'error': 'Test topilmadi'}, status=404)
    except User.DoesNotExist:
        return JsonResponse({'error': 'O\'quvchi topilmadi'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def student_test_management(request):
    """Admin uchun o'quvchilarning test holatlarini boshqarish"""
    if request.user.role != 'admin':
        return redirect('accounts:dashboard')
    
    try:
        # Barcha faol testlar - Raw SQL yordamida olish (is_paused maydonini tekshirmaslik uchun)
        from django.db import connection
        test_ids = []
        try:
            with connection.cursor() as cursor:
                cursor.execute("SELECT id FROM tests_app_test WHERE is_active = 1")
                test_ids = [row[0] for row in cursor.fetchall()]
        except Exception as db_error:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f'Error fetching test IDs: {str(db_error)}')
            # Agar xatolik bo'lsa, bo'sh ro'yxat qaytarish
            test_ids = []
        
        # Test obyektlarini xavfsiz yuklash
        tests = []
        for test_id in test_ids:
            try:
                test = Test.objects.get(id=test_id)
                tests.append(test)
            except Exception:
                continue
        
        # Barcha tasdiqlangan o'quvchilar
        students = User.objects.filter(role='student', is_verified=True)
        
        # Har bir o'quvchi va test uchun urinishlar ma'lumotlari
        student_test_data = []
        
        for student in students:
            student_tests = []
            for test in tests:
                try:
                    attempts = TestAttempt.objects.filter(student=student, test=test)
                    latest_attempt = attempts.order_by('-started_at').first()
                    
                    # Qayta ishlash so'rovlari
                    retake_requests = TestRetakeRequest.objects.filter(
                        student=student,
                        test=test
                    ).order_by('-created_at')
                    
                    test_info = {
                        'test': test,
                        'attempts_count': attempts.count(),
                        'latest_attempt': latest_attempt,
                        'can_retake': latest_attempt is not None,
                        'retake_requests': retake_requests
                    }
                    student_tests.append(test_info)
                except Exception as e:
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.error(f'Error processing test {test.id} for student {student.id}: {str(e)}')
                    continue
            
            student_test_data.append({
                'student': student,
                'tests': student_tests
            })
        
        context = {
            'student_test_data': student_test_data,
            'all_tests': tests
        }
        
        return render(request, 'tests_app/student_test_management.html', context)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Error in student_test_management: {str(e)}', exc_info=True)
        from django.http import HttpResponse
        return HttpResponse(
            f'<html><body><h1>Xatolik</h1><p>Ma\'lumotlarni yuklashda xatolik yuz berdi.</p>'
            f'<p>Migrationni ishga tushiring: <code>python manage.py migrate tests_app</code></p>'
            f'<p>Xatolik: {str(e)}</p></body></html>',
            status=500
        )

@login_required
def edit_test_view(request, test_id):
    """Edit an existing test and its questions (teachers only)"""
    test = get_object_or_404(Test, id=test_id, created_by=request.user)
    if request.user.role != 'teacher':
        return JsonResponse({'error': 'Access denied'}, status=403)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            # Update test fields
            test.title = data.get('title', test.title)
            test.description = data.get('description', test.description)
            test.subject = data.get('subject', test.subject)
            test.grade = int(data.get('grade', test.grade))
            test.time_limit = int(data.get('time_limit', test.time_limit))
            test.max_attempts = int(data.get('max_attempts', test.max_attempts))
            test.show_results = data.get('show_results', test.show_results)
            test.is_active = data.get('is_active', test.is_active)
            test.save()

            # Update questions
            questions_data = data.get('questions', [])
            # Remove old questions not in new data
            new_ids = [q.get('id') for q in questions_data if q.get('id')]
            test.questions.exclude(id__in=new_ids).delete()
            for i, q_data in enumerate(questions_data):
                if q_data.get('id'):
                    # Update existing question
                    question = Question.objects.get(id=q_data['id'], test=test)
                    question.question_text = q_data['question_text']
                    question.question_type = q_data['question_type']
                    question.points = float(q_data.get('points', 1.0))
                    question.order = i + 1
                    question.explanation = q_data.get('explanation', '')
                    question.save()
                    # Update choices
                    if q_data['question_type'] in ['single_choice', 'multiple_choice']:
                        choices_data = q_data.get('choices', [])
                        new_choice_ids = [c.get('id') for c in choices_data if c.get('id')]
                        question.choices.exclude(id__in=new_choice_ids).delete()
                        for c_data in choices_data:
                            if c_data.get('id'):
                                choice = Choice.objects.get(id=c_data['id'], question=question)
                                choice.choice_text = c_data['text']
                                choice.is_correct = c_data.get('is_correct', False)
                                choice.save()
                            else:
                                Choice.objects.create(
                                    question=question,
                                    choice_text=c_data['text'],
                                    is_correct=c_data.get('is_correct', False)
                                )
                    else:
                        question.choices.all().delete()
                else:
                    # Create new question
                    question = Question.objects.create(
                        test=test,
                        question_text=q_data['question_text'],
                        question_type=q_data['question_type'],
                        points=float(q_data.get('points', 1.0)),
                        order=i + 1,
                        explanation=q_data.get('explanation', '')
                    )
                    if q_data['question_type'] in ['single_choice', 'multiple_choice']:
                        for c_data in q_data.get('choices', []):
                            Choice.objects.create(
                                question=question,
                                choice_text=c_data['text'],
                                is_correct=c_data.get('is_correct', False)
                            )
            # Saqlangan savollarni JSON ko‘rinishda qaytarish:
            questions = test.questions.all().order_by('order')
            questions_data = []
            for q in questions:
                q_data = {
                    "question_text": q.question_text,
                    "question_type": q.question_type,
                    "points": q.points,
                    "explanation": q.explanation,
                    "choices": [
                        {"text": c.choice_text, "is_correct": c.is_correct}
                        for c in q.choices.all()
                    ]
                }
                questions_data.append(q_data)
            return JsonResponse({"success": True, "questions": questions_data})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)

    # GET: Render edit page with test and questions
    questions = test.questions.all().order_by('order')
    questions_data = []
    for q in questions:
        q_data = {
            'id': q.id,
            'question_text': q.question_text,
            'question_type': q.question_type,
            'points': q.points,
            'explanation': q.explanation,
            'choices': []
        }
        if q.question_type in ['single_choice', 'multiple_choice']:
            q_data['choices'] = [
                {'id': c.id, 'text': c.choice_text, 'is_correct': c.is_correct}
                for c in q.choices.all()
            ]
        questions_data.append(q_data)
    context = {
        'test': test,
        'questions': questions_data
    }
    return render(request, 'tests_app/edit_test.html', context)

@login_required
def start_test_view(request, test_id):
    """Admin tomonidan o'quvchi uchun testi boshlash"""
    test = get_object_or_404(Test, pk=test_id)
    questions = list(test.questions.all())
    random.shuffle(questions)  # Har bir o‘quvchi uchun random tartib
    
    context = {
        'test': test,
        'questions': questions,
    }
    return render(request, 'tests_app/start_test.html', context)
